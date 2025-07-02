import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

print("🏡 Comprehensive Loan Comparison Calculator with Dynamic Monthly Breakdown")

# Loan program selection
program = input("Enter loan program (Conventional, FHA, VA, USDA): ").strip().lower()

# Purchase price & shared details
amount = float(input("Enter home purchase price (e.g., 400000): "))
close_date_str = input("Enter estimated close date (MM/DD/YYYY): ").strip()
close_date = datetime.datetime.strptime(close_date_str, "%m/%d/%Y")
last_day = datetime.datetime(close_date.year, close_date.month, 28) + datetime.timedelta(days=4)
last_day -= datetime.timedelta(days=last_day.day)
interim_days = (last_day - close_date).days + 1

borrower_name = input("Enter borrower name: ").strip()
credit_score = int(input("Enter borrower credit score: "))

escrow_waived = input("Are escrows waived? (yes/no): ").strip().lower() == "yes"
taxes = float(input("Estimated monthly property taxes: "))
insurance = float(input("Estimated monthly homeowners insurance: "))

# Down payment options
downs = []
if program == "fha":
    num_downs = int(input("How many down payment options? (min 3.5%): "))
    for i in range(1, num_downs + 1):
        down = float(input(f"Down payment option {i} (%): "))
        if down < 3.5:
            print("❌ FHA requires at least 3.5% down.")
            continue
        downs.append(down)
elif program == "va":
    num_downs = int(input("How many down payment options? (VA allows 0%+): "))
    for i in range(1, num_downs + 1):
        down = float(input(f"Down payment option {i} (%): "))
        downs.append(down)
elif program == "usda":
    num_downs = int(input("How many down payment options? (USDA allows 0%+): "))
    for i in range(1, num_downs + 1):
        down = float(input(f"Down payment option {i} (%): "))
        downs.append(down)
else:  # Conventional
    first_time = input("Is the borrower a first-time homebuyer? (yes/no): ").strip().lower() == "yes"
    num_downs = int(input("How many down payment options? "))
    for i in range(1, num_downs + 1):
        down = float(input(f"Down payment option {i} (%): "))
        if first_time and down < 3:
            print("❌ First-time conventional requires ≥3% down.")
            continue
        if not first_time and down < 5:
            print("❌ Non-first-time conventional requires ≥5% down.")
            continue
        downs.append(down)

# Rate options & lender credits
num_rates = int(input("How many interest rate options? "))
rates = [float(input(f"Interest rate option {i} (%): ")) for i in range(1, num_rates + 1)]
credits = [float(input(f"Estimated lender credit for rate option {i}: ")) for i in range(1, num_rates + 1)]

# VA-specific
if program == "va":
    exempt_fee = input("Is borrower exempt from VA funding fee? (yes/no): ").strip().lower() == "yes"
    first_use = input("Is this the borrower's first use of VA? (yes/no): ").strip().lower() == "yes"

# Results
results = []
for down in downs:
    down_amt = amount * (down / 100)
    base_loan = amount - down_amt
    ltv = base_loan / amount * 100

    for idx, rate in enumerate(rates):
        loan_amt = base_loan
        monthly_mi = 0
        uf_fee = 0

        # PMI/MIP/Funding Fee
        if program == "fha":
            uf_fee = base_loan * 0.0175
            loan_amt += uf_fee
            annual_mip = 0.0055 if ltv >= 90 else 0.0050
            monthly_mi = (loan_amt * annual_mip) / 12
        elif program == "va" and not exempt_fee:
            if first_use:
                uf_fee = base_loan * (0.0215 if down < 5 else 0.015 if down < 10 else 0.0125)
            else:
                uf_fee = base_loan * (0.033 if down < 5 else 0.015 if down < 10 else 0.0125)
            loan_amt += uf_fee
        elif program == "usda":
            uf_fee = base_loan * 0.01
            loan_amt += uf_fee
            monthly_mi = (loan_amt * 0.0035) / 12
        elif program == "conventional" and ltv > 80:
            pmi_factor = 0.0062
            if credit_score >= 740:
                pmi_factor *= 0.85
            monthly_mi = (loan_amt * pmi_factor) / 12

        # P&I calc
        n, r = 30 * 12, rate / 100 / 12
        pi = loan_amt * (r * (1 + r) ** n) / ((1 + r) ** n - 1)
        monthly_escrow = 0 if escrow_waived else (taxes + insurance)
        total_payment = pi + monthly_mi + monthly_escrow

        interim_int = ((loan_amt * rate / 100) / 12 / 30) * interim_days
        homeowners_prem = insurance * 12
        prop_tax_res = taxes * 6
        prepaids = interim_int + homeowners_prem + prop_tax_res

        closing_costs = {
            "Credit Report": 120,
            "Underwriter Fee": 1250,
            "Flood/Tax Cert": 76,
            "Appraisal Fee": 625,
            "Lenders Title Insurance": 985.20,
            "Closing Attorney": 925,
            "Recording Fee": 351.86,
        }

        cash_close = down_amt + sum(closing_costs.values()) + prepaids - credits[idx]

        results.append({
            "Down %": down,
            "Rate": rate,
            "Loan Amount": loan_amt,
            "LTV %": ltv,
            "Monthly P&I": pi,
            "Monthly MI": monthly_mi,
            "Taxes": taxes if not escrow_waived else 0,
            "Insurance": insurance if not escrow_waived else 0,
            "Total Payment": total_payment,
            "Lender Credit": credits[idx],
            "Cash to Close": cash_close,
            "Interim Interest": interim_int,
            "Homeowners Premium": homeowners_prem,
            "Property Tax Reserve": prop_tax_res,
            "Prepaids Total": prepaids,
        })

# Excel export
wb = Workbook()
ws = wb.active
ws.title = "Loan Comparison"

bold_font = Font(bold=True)
yellow_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
currency_fmt = '"$"#,##0.00'
percent_fmt = '0.000%'

# Fixed labels (they start the sheet and closing costs sections)
labels = [
    "Loan Term (years)",
    "Purchase Price",
    "Loan Amount",
    "Mortgage Rate",
    "Monthly P&I",
]

# Write fixed labels
for row, label in enumerate(labels, start=1):
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(bold=True)

# Loop through options to fill columns
for opt_idx, option in enumerate(results, start=2):
    col = opt_idx
    ws.cell(row=1, column=col, value=f"Option {opt_idx-1}").font = bold_font
    ws.cell(row=2, column=col, value=30)
    ws.cell(row=3, column=col, value=amount).number_format = currency_fmt
    ws.cell(row=4, column=col, value=option["Loan Amount"]).number_format = currency_fmt
    ws.cell(row=5, column=col, value=option["Rate"] / 100).number_format = percent_fmt
    ws.cell(row=6, column=col, value=option["Monthly P&I"]).number_format = currency_fmt

    current_row = 7

    # Dynamic Monthly MI section
    include_mi = option["Monthly MI"] > 0
    mi_label = ""
    if include_mi:
        if program == "fha":
            mi_label = "Monthly MIP"
        elif program == "usda":
            mi_label = "Monthly USDA Fee"
        else:
            mi_label = "Monthly PMI"
        ws.cell(row=current_row, column=1, value=mi_label).font = Font(bold=True)
        ws.cell(row=current_row, column=col, value=option["Monthly MI"]).number_format = currency_fmt
        current_row += 1

    if not escrow_waived:
        ws.cell(row=current_row, column=1, value="Taxes").font = Font(bold=True)
        ws.cell(row=current_row, column=col, value=option["Taxes"]).number_format = currency_fmt
        current_row += 1

        ws.cell(row=current_row, column=1, value="Insurance").font = Font(bold=True)
        ws.cell(row=current_row, column=col, value=option["Insurance"]).number_format = currency_fmt
        current_row += 1

    # Closing Costs header
    current_row += 1
    ws.cell(row=current_row, column=1, value="Closing Costs").font = bold_font
    ws.cell(row=current_row, column=1).fill = yellow_fill
    current_row += 1

    ws.cell(row=current_row, column=1, value="Lender Credit").font = Font(bold=True)
    ws.cell(row=current_row, column=col, value=option["Lender Credit"]).number_format = currency_fmt
    current_row += 1

    for k, v in {
        "Credit Report": 120,
        "Underwriter Fee": 1250,
        "Flood/Tax Cert": 76,
        "Appraisal Fee": 625,
        "Lenders Title Insurance": 985.20,
        "Closing Attorney": 925,
        "Recording Fee": 351.86,
    }.items():
        ws.cell(row=current_row, column=1, value=k).font = Font(bold=True)
        ws.cell(row=current_row, column=col, value=v).number_format = currency_fmt
        current_row += 1

    # Pre-paid Items header
    current_row += 1
    ws.cell(row=current_row, column=1, value="Pre-paid Items").font = bold_font
    ws.cell(row=current_row, column=1).fill = yellow_fill
    current_row += 1

    ws.cell(row=current_row, column=1, value="Interim Interest").font = Font(bold=True)
    ws.cell(row=current_row, column=col, value=option["Interim Interest"]).number_format = currency_fmt
    current_row += 1

    ws.cell(row=current_row, column=1, value="Homeowners Insurance Premium").font = Font(bold=True)
    ws.cell(row=current_row, column=col, value=option["Homeowners Premium"]).number_format = currency_fmt
    current_row += 1

    ws.cell(row=current_row, column=1, value="Property Tax Reserve").font = Font(bold=True)
    ws.cell(row=current_row, column=col, value=option["Property Tax Reserve"]).number_format = currency_fmt
    current_row += 1

    ws.cell(row=current_row, column=1, value="Total Prepaid Items").font = Font(bold=True)
    ws.cell(row=current_row, column=col, value=option["Prepaids Total"]).number_format = currency_fmt
    current_row += 2

    # Transaction Summary header
    ws.cell(row=current_row, column=1, value="Transaction Summary").font = bold_font
    ws.cell(row=current_row, column=1).fill = yellow_fill
    current_row += 1

    ws.cell(row=current_row, column=1, value="Down Payment").font = Font(bold=True)
    ws.cell(row=current_row, column=col, value=amount * (option["Down %"] / 100)).number_format = currency_fmt
    current_row += 1

    ws.cell(row=current_row, column=1, value="Total Closing Costs").font = Font(bold=True)
    ws.cell(row=current_row, column=col, value=sum([120,1250,76,625,985.20,925,351.86])).number_format = currency_fmt
    current_row += 1

    ws.cell(row=current_row, column=1, value="Total Prepaid Items").font = Font(bold=True)
    ws.cell(row=current_row, column=col, value=option["Prepaids Total"]).number_format = currency_fmt
    current_row += 1

    ws.cell(row=current_row, column=1, value="Total Cash Due at Closing").font = Font(bold=True)
    ws.cell(row=current_row, column=col, value=option["Cash to Close"]).number_format = currency_fmt

# Auto-fit columns
for col in ws.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_len + 2

filename = f"{borrower_name.replace(' ', '_')}_loan_options.xlsx"
wb.save(filename)
print(f"\n✅ Table exported as '{filename}'")





