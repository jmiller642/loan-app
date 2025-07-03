import streamlit as st
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

st.title("🏡 Comprehensive Loan Calculator with Closing Costs and Comparison Table")

# Borrower info
borrower_name = st.text_input("Borrower name")
credit_score = st.number_input("Borrower credit score", min_value=300, max_value=850, value=740)

# Loan details
program = st.selectbox("Loan program", ["Conventional", "FHA", "VA", "USDA"])
amount = st.number_input("Home purchase price ($)", min_value=10000.0, value=300000.0, step=1000.0)
close_date = st.date_input("Estimated close date", datetime.date.today())

escrow_waived = st.checkbox("Escrows waived?")
taxes = st.number_input("Estimated monthly property taxes ($)", min_value=0.0, value=300.0, step=10.0)
insurance = st.number_input("Estimated monthly homeowners insurance ($)", min_value=0.0, value=100.0, step=10.0)

last_day = datetime.date(close_date.year, close_date.month, 28) + datetime.timedelta(days=4)
last_day -= datetime.timedelta(days=last_day.day)
interim_days = (last_day - close_date).days + 1

# Down payment options
downs = []
if program == "FHA":
    num_downs = st.number_input("How many down payment options? (min 3.5%)", min_value=1, step=1, value=1)
    for i in range(num_downs):
        down = st.number_input(f"Down payment option {i+1} (%)", min_value=3.5, step=0.1)
        downs.append(down)
elif program == "VA":
    num_downs = st.number_input("How many down payment options? (VA allows 0%+)", min_value=1, step=1, value=1)
    for i in range(num_downs):
        down = st.number_input(f"Down payment option {i+1} (%)", min_value=0.0, step=0.1)
        downs.append(down)
elif program == "USDA":
    num_downs = st.number_input("How many down payment options? (USDA allows 0%+)", min_value=1, step=1, value=1)
    for i in range(num_downs):
        down = st.number_input(f"Down payment option {i+1} (%)", min_value=0.0, step=0.1)
        downs.append(down)
else:  # Conventional
    first_time = st.checkbox("First-time homebuyer?")
    num_downs = st.number_input("How many down payment options?", min_value=1, step=1, value=1)
    min_down = 3 if first_time else 5
    for i in range(num_downs):
        down = st.number_input(f"Down payment option {i+1} (%) (min {min_down}%)", min_value=min_down, step=0.1)
        downs.append(down)

num_rates = st.number_input("How many interest rate options?", min_value=1, step=1, value=1)
rates, credits = [], []
for i in range(num_rates):
    rate = st.number_input(f"Interest rate option {i+1} (%)", min_value=0.1, step=0.01)
    credit = st.number_input(f"Lender credit for rate option {i+1} ($)", step=100.0)
    rates.append(rate)
    credits.append(credit)

# VA-specific inputs
exempt_fee, first_use = False, True
if program == "VA":
    exempt_fee = st.checkbox("Borrower exempt from VA funding fee?")
    first_use = st.checkbox("Is this the borrower's first use of VA?", value=True)

if st.button("Generate Loan Options Table"):
    results = []
    for down in downs:
        down_amt = amount * (down / 100)
        base_loan = amount - down_amt
        ltv = base_loan / amount * 100

        for idx, rate in enumerate(rates):
            loan_amt = base_loan
            monthly_mi, uf_fee = 0, 0

            # PMI/MIP/Funding Fee
            if program == "FHA":
                uf_fee = base_loan * 0.0175
                loan_amt += uf_fee
                annual_mip = 0.0055 if ltv >= 90 else 0.0050
                monthly_mi = (loan_amt * annual_mip) / 12
            elif program == "VA" and not exempt_fee:
                uf_fee = base_loan * (0.0215 if first_use and down < 5 else 0.015 if first_use else 0.033)
                loan_amt += uf_fee
            elif program == "USDA":
                uf_fee = base_loan * 0.01
                loan_amt += uf_fee
                monthly_mi = (loan_amt * 0.0035) / 12
            elif program == "Conventional" and ltv > 80:
                pmi_factor = 0.0062
                if credit_score >= 740:
                    pmi_factor *= 0.85
                monthly_mi = (loan_amt * pmi_factor) / 12

            # Monthly P&I calculation
            n, r_monthly = 30 * 12, rate / 100 / 12
            pi = loan_amt * (r_monthly * (1 + r_monthly) ** n) / ((1 + r_monthly) ** n - 1)
            monthly_escrow = 0 if escrow_waived else (taxes + insurance)
            total_payment = pi + monthly_mi + monthly_escrow

            interim_int = ((loan_amt * rate / 100) / 12 / 30) * interim_days
            homeowners_prem = insurance * 12
            prop_tax_res = taxes * 6
            prepaids = interim_int + homeowners_prem + prop_tax_res

            closing_costs = {
                "Credit Report": 120,
                "Underwriter Fee": 1250,
                "Flood/Tax Cert": 76,
                "Appraisal Fee": 625,
                "Lenders Title Insurance": 985.20,
                "Closing Attorney": 925,
                "Recording Fee": 351.86,
            }

            cash_close = down_amt + sum(closing_costs.values()) + prepaids - credits[idx]

            results.append({
                "Down %": down,
                "Rate": rate,
                "Loan Amount": loan_amt,
                "Monthly P&I": pi,
                "Monthly MI/MIP/USDA": monthly_mi,
                "Taxes": taxes if not escrow_waived else 0,
                "Insurance": insurance if not escrow_waived else 0,
                "Total Payment": total_payment,
                "Lender Credit": credits[idx],
                "Cash to Close": cash_close,
            })

    # Show table in Streamlit
    st.write("## Loan Options Comparison")
    st.dataframe(results)

    # Excel export
    wb = Workbook()
    ws = wb.active
    ws.title = "Loan Comparison"
    bold = Font(bold=True)
    currency_fmt = '"$"#,##0.00'
    percent_fmt = '0.000%'

    labels = [
        "Down Payment %", "Rate", "Loan Amount", "Monthly P&I",
        "Monthly MI/MIP/USDA", "Taxes", "Insurance", "Total Payment",
        "Lender Credit", "Cash to Close"
    ]

    for row, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = bold

    for col_idx, option in enumerate(results, start=2):
        ws.cell(row=1, column=col_idx, value=f"Option {col_idx-1}").font = bold
        ws.cell(row=2, column=col_idx, value=option["Down %"]/100).number_format = percent_fmt
        ws.cell(row=3, column=col_idx, value=option["Rate"]/100).number_format = percent_fmt
        ws.cell(row=4, column=col_idx, value=option["Loan Amount"]).number_format = currency_fmt
        ws.cell(row=5, column=col_idx, value=option["Monthly P&I"]).number_format = currency_fmt
        ws.cell(row=6, column=col_idx, value=option["Monthly MI/MIP/USDA"]).number_format = currency_fmt
        ws.cell(row=7, column=col_idx, value=option["Taxes"]).number_format = currency_fmt
        ws.cell(row=8, column=col_idx, value=option["Insurance"]).number_format = currency_fmt
        ws.cell(row=9, column=col_idx, value=option["Total Payment"]).number_format = currency_fmt
        ws.cell(row=10, column=col_idx, value=option["Lender Credit"]).number_format = currency_fmt
        ws.cell(row=11, column=col_idx, value=option["Cash to Close"]).number_format = currency_fmt

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    filename = f"{borrower_name.replace(' ', '_')}_loan_options.xlsx"
    wb.save(filename)
    st.success(f"✅ Table exported as '{filename}'")

